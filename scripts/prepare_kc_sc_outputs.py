#!/usr/bin/env python3
"""
Prepare KC and SC AE output folders to match the NCC deliverable pattern.

This script starts from the already-produced AE CSVs, adds case-file container
Barcode/Location ID values by accepted name-range container lists, appends the
28 KC adult records missing from the current Adults source, and writes paired
review CSVs for unresolved lookup cases.
"""

from __future__ import annotations

import csv
import re
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import migrate


CASE_FILE_CONTAINER_TYPES = {"Legal Upright - 5"}
KC_BASE_ARCHIVE = ROOT / "Orphan's Court -_ AE_ KC" / "Orphans.zip"
KC_BASE_OUTPUT_MEMBER = "Orphans/Orphan's Court -_ AE_ KC/kc_output_children_single.csv"
KC_MISSING_ADULTS_CANDIDATES = [
    ROOT / "Orphan's Court -_ AE_ KC" / "KC Copy of Adults Missing from Adults.xlsx",
    ROOT / "Orphan's Court -_ AE_ KC" / "Records Missing from Adults.xlsx",
]
OVERSIZE_SOURCE_NOTE = "Oversize source barcode; letter-size companion barcode: "
LETTER_SIZE_COMPANION_NOTE = "Letter-size companion entry; source oversize barcode: "
SOURCE_BARCODE_IS_RANGE_NOTE = "Source barcode already matches letter-size range"


@dataclass(frozen=True)
class ContainerRange:
    start: tuple[str, str]
    end: tuple[str, str]
    barcode: str
    location_id: str
    title: str


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        try:
            return str(int(float(text)))
        except ValueError:
            return text
    return text


def name_key(last_name: str, first_name: str = "") -> tuple[str, str]:
    return (last_name.casefold().strip(), first_name.casefold().strip())


def parse_name(text: str) -> tuple[str, str]:
    text = text.strip().lstrip("? ")
    if "," not in text:
        return (text, "")
    last_name, first_name = text.split(",", 1)
    return (last_name.strip(), first_name.strip())


def parse_container_title(title: str) -> tuple[tuple[str, str], tuple[str, str]] | None:
    title = title.strip()
    parts = re.split(r"\s+(?:to|thru|through)\s+", title, maxsplit=1, flags=re.I)
    if len(parts) == 2:
        return parse_name(parts[0]), parse_name(parts[1])

    parts = re.split(r"\s+-\s+", title, maxsplit=1)
    if len(parts) == 2:
        return parse_name(parts[0]), parse_name(parts[1])

    if "," in title:
        parsed = parse_name(title)
        return parsed, parsed

    if title and len(title.split()) == 1:
        return (title, ""), (title, "zzzz")

    return None


def load_case_file_ranges(path: Path) -> list[ContainerRange]:
    ranges: list[ContainerRange] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            container_type = clean_text(row.get("Container Type"))
            if container_type not in CASE_FILE_CONTAINER_TYPES:
                continue

            title = clean_text(row.get("Title"))
            parsed = parse_container_title(title)
            if parsed is None:
                continue

            start, end = parsed
            ranges.append(
                ContainerRange(
                    start=name_key(*start),
                    end=name_key(*end),
                    barcode=clean_text(row.get("Barcode")),
                    location_id=clean_text(row.get("Location ID")),
                    title=title,
                )
            )

    ranges.sort(key=lambda item: item.start)
    return ranges


def load_barcode_locations(path: Path) -> dict[str, str]:
    locations: dict[str, str] = {}
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            barcode = clean_text(row.get("Barcode"))
            location_id = clean_text(row.get("Location ID"))
            if barcode and location_id:
                locations[barcode] = location_id
    return locations


def find_container(
    ranges: list[ContainerRange],
    last_name: str,
    first_name: str = "",
) -> ContainerRange | None:
    if not last_name:
        return None

    lookup = name_key(last_name, first_name)
    for item in ranges:
        if item.start <= lookup <= item.end:
            return item

    last_only = name_key(last_name)
    for item in ranges:
        if item.start[0] <= last_only[0] <= item.end[0]:
            return item

    return None


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))
    if not rows:
        raise ValueError(f"{path} is empty")
    header = rows[0]
    data = [row + [""] * (len(header) - len(row)) for row in rows[1:]]
    return header, [row[: len(header)] for row in data]


def write_csv(path: Path, header: list[str], rows: list[list[str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        writer.writerows(rows)


def first_index(header: list[str], column: str) -> int:
    return header.index(column)


def column_indices(header: list[str], column: str) -> list[int]:
    return [idx for idx, name in enumerate(header) if name == column]


def append_note(row: list[str], note_indices: list[int], note: str) -> None:
    if not note_indices:
        return

    target_idx = note_indices[1] if len(note_indices) > 1 else note_indices[0]
    current = clean_text(row[target_idx])
    if note in current:
        return
    row[target_idx] = f"{current} | {note}" if current else note


def has_note(row: list[str], note_indices: list[int], fragment: str) -> bool:
    return any(fragment in clean_text(row[idx]) for idx in note_indices)


def is_oversize_row(row: list[str], note_indices: list[int]) -> bool:
    if not note_indices:
        return False
    return "oversize: yes" in clean_text(row[note_indices[0]]).casefold()


def infer_first_from_title(title: str, last_name: str) -> str:
    title = title.strip()
    last_name = last_name.strip()
    if not title or not last_name:
        return ""

    if not title.casefold().endswith(last_name.casefold()):
        return ""

    prefix = title[: -len(last_name)].strip()
    if not prefix or " and " in prefix.casefold():
        return ""

    return prefix.split()[0].strip(",")


def apply_container_lookup(
    county: str,
    header: list[str],
    rows: list[list[str]],
    ranges: list[ContainerRange],
    barcode_locations: dict[str, str],
    source_barcodes: dict[str, str] | None = None,
    add_oversize_source_companions: bool = False,
) -> list[str]:
    title_idx = first_index(header, "Title")
    record_id_idx = first_index(header, "Record_ID")
    last_name_idx = first_index(header, "Deceaseds_ Last_name")
    barcode_idx = first_index(header, "Barcode")
    location_idx = first_index(header, "Location ID")
    note_indices = column_indices(header, "Notes")
    source_barcodes = source_barcodes or {}
    companion_rows: list[list[str]] = []
    flags: list[str] = []

    for row in rows:
        record_id = clean_text(row[record_id_idx])
        last_name = clean_text(row[last_name_idx])
        first_name = infer_first_from_title(clean_text(row[title_idx]), last_name)
        existing_barcode = clean_text(row[barcode_idx])
        source_barcode = source_barcodes.get(record_id, "")

        if existing_barcode:
            if existing_barcode in barcode_locations:
                row[location_idx] = barcode_locations[existing_barcode]
            elif not clean_text(row[location_idx]):
                flags.append(
                    f"{county} record {record_id}: Existing barcode "
                    f"{existing_barcode} not found in container list"
                )

            if (
                add_oversize_source_companions
                and source_barcode
                and source_barcode == existing_barcode
                and is_oversize_row(row, note_indices)
                and not has_note(row, note_indices, LETTER_SIZE_COMPANION_NOTE)
            ):
                if not last_name:
                    flags.append(
                        f"{county} record {record_id}: Source oversize barcode "
                        "present, but no last name available for letter-size companion lookup"
                    )
                    continue

                match = find_container(ranges, last_name, first_name)
                if match is None:
                    flags.append(
                        f"{county} record {record_id}: Source oversize barcode "
                        f"{existing_barcode} present, but no letter-size container match for '{last_name}'"
                    )
                    continue

                if match.barcode == existing_barcode:
                    append_note(row, note_indices, SOURCE_BARCODE_IS_RANGE_NOTE)
                    flags.append(
                        f"{county} record {record_id}: Source barcode {existing_barcode} "
                        "already matches the letter-size range; no companion row added"
                    )
                    continue

                companion_exists = any(
                    clean_text(candidate[record_id_idx]) == record_id
                    and clean_text(candidate[barcode_idx]) == match.barcode
                    and has_note(candidate, note_indices, LETTER_SIZE_COMPANION_NOTE)
                    for candidate in rows + companion_rows
                )
                if companion_exists:
                    continue

                companion = row.copy()
                append_note(row, note_indices, f"{OVERSIZE_SOURCE_NOTE}{match.barcode}")
                companion[barcode_idx] = match.barcode
                companion[location_idx] = match.location_id
                append_note(companion, note_indices, f"{LETTER_SIZE_COMPANION_NOTE}{existing_barcode}")
                companion_rows.append(companion)
                flags.append(
                    f"{county} record {record_id}: Added letter-size companion "
                    f"barcode {match.barcode} for source oversize barcode {existing_barcode}"
                )
            continue

        if not last_name:
            flags.append(f"{county} record {record_id}: No last name available for container lookup")
            continue

        match = find_container(ranges, last_name, first_name)
        if match is None:
            flags.append(f"{county} record {record_id}: No container match for '{last_name}'")
            continue

        row[barcode_idx] = match.barcode
        row[location_idx] = match.location_id

    rows.extend(companion_rows)
    return flags


def review_path_for(output_path: Path) -> Path:
    return output_path.with_name(f"{output_path.stem}_review{output_path.suffix}")


def write_review(path: Path, flags: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Flag"])
        for flag in flags:
            writer.writerow([flag])


def output_row_from_parts(
    header: list[str],
    output: dict[str, str],
    description_values: list[str],
    notes_values: list[str],
) -> list[str]:
    row: list[str] = []
    description_index = 0
    notes_index = 0
    for column in header:
        if column == "Description":
            value = ""
            if description_index < len(description_values):
                value = description_values[description_index]
            row.append(value)
            description_index += 1
        elif column == "Notes":
            value = ""
            if notes_index < len(notes_values):
                value = notes_values[notes_index]
            row.append(value)
            notes_index += 1
        else:
            row.append(output.get(column, ""))
    return row


def load_kc_missing_adults(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [clean_text(value) for value in rows[0]]
    adults: list[dict[str, str]] = []
    for row in rows[1:]:
        adult = {column: clean_text(value) for column, value in zip(header, row)}
        if adult.get("Record ID"):
            adults.append(adult)
    return adults


def choose_existing_path(paths: list[Path]) -> Path:
    for path in paths:
        if path.exists():
            return path
    raise FileNotFoundError(f"None of these paths exist: {paths}")


def load_source_barcodes_from_output_archive(
    archive_path: Path,
    member_name: str,
) -> dict[str, str]:
    if not archive_path.exists():
        return {}

    with zipfile.ZipFile(archive_path) as archive:
        with archive.open(member_name) as handle:
            lines = (line.decode("utf-8-sig") for line in handle)
            reader = csv.reader(lines)
            header = next(reader)
            record_id_idx = first_index(header, "Record_ID")
            barcode_idx = first_index(header, "Barcode")
            source_barcodes: dict[str, str] = {}
            for row in reader:
                row = row + [""] * (len(header) - len(row))
                record_id = clean_text(row[record_id_idx])
                barcode = clean_text(row[barcode_idx])
                if record_id and barcode:
                    source_barcodes[record_id] = barcode
            return source_barcodes


def kc_missing_adult_barcodes(missing_adults: list[dict[str, str]]) -> dict[str, str]:
    source_barcodes: dict[str, str] = {}
    for adult in missing_adults:
        record_id = clean_text(adult.get("Record ID"))
        barcode = clean_text(adult.get("Barcode"))
        if record_id and barcode:
            source_barcodes[record_id] = barcode
    return source_barcodes


def append_kc_missing_rows(
    header: list[str],
    rows: list[list[str]],
    missing_adults: list[dict[str, str]],
    children_format: str,
) -> int:
    record_id_idx = first_index(header, "Record_ID")
    existing_ids = {clean_text(row[record_id_idx]) for row in rows}
    description_count = len(column_indices(header, "Description"))
    notes_count = len(column_indices(header, "Notes"))

    migrate.PROFILE = "kc"
    migrate.ADULT_COLUMNS = migrate.ADULT_COLUMNS_KC
    defaults = {
        "rg": "3840",
        "series": "006",
        "subgr": "000",
        "dept_org": "Orphans Court, Kent County",
        "series_name": "Case Files",
    }

    appended = 0
    for adult in missing_adults:
        record_id = clean_text(adult.get("Record ID"))
        if record_id in existing_ids:
            continue

        output, descriptions, notes, _flags = migrate.build_record(
            adult,
            [],
            None,
            defaults,
            children_format=children_format,
            description_count=description_count,
        )
        rows.append(output_row_from_parts(header, output, descriptions, notes[:notes_count]))
        existing_ids.add(record_id)
        appended += 1

    rows.sort(key=lambda row: int(row[record_id_idx]) if row[record_id_idx].isdigit() else 0)
    return appended


def sort_rows_by_record_id(header: list[str], rows: list[list[str]]) -> None:
    record_id_idx = first_index(header, "Record_ID")
    note_indices = column_indices(header, "Notes")

    def sort_key(row: list[str]) -> tuple[int, str, int]:
        record_id = clean_text(row[record_id_idx])
        numeric_id = int(record_id) if record_id.isdigit() else sys.maxsize
        is_companion = 1 if has_note(row, note_indices, LETTER_SIZE_COMPANION_NOTE) else 0
        return numeric_id, record_id, is_companion

    rows.sort(key=sort_key)


def prepare_county(
    county: str,
    output_paths: list[tuple[Path, str]],
    container_path: Path,
    missing_adults_path: Path | None = None,
    source_barcodes: dict[str, str] | None = None,
    add_oversize_source_companions: bool = False,
) -> None:
    ranges = load_case_file_ranges(container_path)
    barcode_locations = load_barcode_locations(container_path)
    missing_adults = load_kc_missing_adults(missing_adults_path) if missing_adults_path else []
    print(
        f"{county}: loaded {len(ranges)} case-file container ranges "
        f"and {len(barcode_locations)} barcode locations"
    )

    for output_path, children_format in output_paths:
        header, rows = read_csv(output_path)
        appended = 0
        if missing_adults:
            appended = append_kc_missing_rows(header, rows, missing_adults, children_format)
        flags = apply_container_lookup(
            county,
            header,
            rows,
            ranges,
            barcode_locations,
            source_barcodes=source_barcodes,
            add_oversize_source_companions=add_oversize_source_companions,
        )
        sort_rows_by_record_id(header, rows)
        write_csv(output_path, header, rows)
        write_review(review_path_for(output_path), flags)
        located = sum(1 for row in rows if row[first_index(header, "Location ID")].strip())
        print(
            f"{county}: {output_path.name} rows={len(rows)} "
            f"locations={located} appended={appended} review_flags={len(flags)}"
        )


def main() -> None:
    kc_missing_adults_path = choose_existing_path(KC_MISSING_ADULTS_CANDIDATES)
    kc_missing_adults = load_kc_missing_adults(kc_missing_adults_path)
    kc_source_barcodes = load_source_barcodes_from_output_archive(
        KC_BASE_ARCHIVE,
        KC_BASE_OUTPUT_MEMBER,
    )
    kc_source_barcodes.update(kc_missing_adult_barcodes(kc_missing_adults))

    prepare_county(
        "KC",
        [
            (ROOT / "Orphan's Court -_ AE_ KC" / "kc_output_children_single.csv", "single"),
            (ROOT / "Orphan's Court -_ AE_ KC" / "kc_output_children_columns.csv", "columns"),
        ],
        ROOT / "Orphan's Court -_ AE_ KC" / "KC Container List.csv",
        kc_missing_adults_path,
        source_barcodes=kc_source_barcodes,
        add_oversize_source_companions=True,
    )
    prepare_county(
        "SC",
        [
            (ROOT / "Orphan's Court -_ AE_ SC" / "sc_output_children_single.csv", "single"),
            (ROOT / "Orphan's Court -_ AE_ SC" / "sc_output_children_columns.csv", "columns"),
        ],
        ROOT / "Orphan's Court -_ AE_ SC" / "SC Container List.csv",
    )


if __name__ == "__main__":
    main()
